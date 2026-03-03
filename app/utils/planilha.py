import io
import csv
from typing import Dict, Any
from datetime import datetime, timezone
from openpyxl import load_workbook
from email_validator import validate_email, EmailNotValidError
from bson import ObjectId

from app.utils.validations import validate_cpf, normalize_participante_data


async def process_planilha(file_bytes: bytes, filename: str, evento_id: str, db, import_id: str = None, validate_only: bool = False) -> Dict[str, Any]:
    """Processa uma planilha (.xlsx ou .csv) e importa participantes/ingressos.

    Retorna um relatório com estatísticas e lista de erros por linha.
    """
    errors = []
    created_participants = 0
    reused_participants = 0
    created_ingressos = 0
    total = 0

    # Carrega evento para obter campos obrigatorios
    evento = await _fetch_evento(db, evento_id)
    if not evento:
        raise ValueError("Evento não encontrado para importação")
    campos_obrigatorios = evento.get("campos_obrigatorios_planilha", ["Nome", "Email", "CPF"])

    # Determinar tipo de arquivo
    if filename.lower().endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"total": 0, "created_participants": 0, "created_ingressos": 0, "errors": []}
        header = [str(c).strip() if c is not None else '' for c in rows[0]]
        data_rows = rows[1:]
        def iter_rows():
            for r in data_rows:
                yield {h: (v if v is not None else '') for h, v in zip(header, r)}
    else:
        # assume csv
        text = file_bytes.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        def iter_rows():
            for r in reader:
                yield r

    seen_cpfs = set()

    line_no = 1
    for row in iter_rows():
        line_no += 1
        
        # Normalize header keys to lowercase without spaces for lookup
        def g(key):
            return row.get(key) or row.get(key.title()) or row.get(key.capitalize()) or row.get(key.upper()) or row.get(key.lower())

        # Extract basic fields
        nome = None
        email = None
        cpf_raw = None
        tipo_num = None
        # Try common header names
        for k in row.keys():
            kn = k.strip().lower()
            if kn in ['nome', 'name']:
                nome = row[k]
            elif kn in ['email', 'e-mail']:
                email = row[k]
            elif kn in ['cpf']:
                # Excel pode interpretar CPF como número
                raw_value = row[k]
                if raw_value is not None:
                    cpf_raw = str(raw_value).strip()
            elif kn in ['tipo ingresso', 'tipo_ingresso', 'tipo', 'tipoingresso']:
                tipo_num = row[k]
        
        # Pula linhas completamente vazias (ignora fórmulas do Excel)
        # Uma linha é considerada vazia se Nome, Email e CPF estão vazios ou são apenas espaços
        is_empty_line = True
        if nome and str(nome).strip():
            is_empty_line = False
        if email and str(email).strip():
            is_empty_line = False
        if cpf_raw and str(cpf_raw).strip() and not str(cpf_raw).startswith('='):
            is_empty_line = False
        
        if is_empty_line:
            continue  # Pula esta linha sem incrementar total ou gerar erros
        
        total += 1
        # After each row processed, update progress if import_id provided
        if import_id:
            try:
                await db.planilha_importacoes.update_one({'_id': ObjectId(import_id)}, {'$set': {'progress': {'processed': total}}})
            except Exception:
                # ignore DB/update issues (e.g., fake DB in tests)
                pass

        row_errors = []
        # Required fields
        for req in campos_obrigatorios:
            key = req.strip().lower()
            if key == 'nome' and not nome:
                row_errors.append('Nome obrigatório')
            if key == 'email' and not email:
                row_errors.append('Email obrigatório')
            if key == 'cpf' and not cpf_raw:
                row_errors.append('CPF obrigatório')

        # CPF validation
        cpf_digits = None
        if cpf_raw:
            try:
                cpf_digits = validate_cpf(str(cpf_raw))
            except ValueError as e:
                row_errors.append(f'CPF inválido: {e}')
        # Email validation (relaxed): must contain '@' and a domain with a dot
        if email:
            try:
                email_s = str(email)
                parts = email_s.split('@')
                if len(parts) != 2 or '.' not in parts[1]:
                    row_errors.append('Email inválido')
            except Exception:
                row_errors.append('Email inválido')
        # Tipo ingresso validation
        tipo_obj = None
        if tipo_num:
            try:
                numero = int(str(tipo_num).strip())
                tipo_obj = await db.tipos_ingresso.find_one({'evento_id': evento_id, 'numero': numero})
                if not tipo_obj:
                    row_errors.append('Tipo de ingresso inválido para o evento')
            except Exception:
                row_errors.append('Tipo de ingresso deve ser número inteiro')

        # Dup CPF in sheet
        if cpf_digits:
            if cpf_digits in seen_cpfs:
                row_errors.append('CPF duplicado na planilha')
            else:
                seen_cpfs.add(cpf_digits)

        # Dup CPF in DB: check participantes then ingressos
        if cpf_digits and not row_errors:
            existing_part = await db.participantes.find_one({'cpf': cpf_digits})
            if existing_part:
                # Verifica se participante já tem ingresso para este evento (procura em ingressos embutidos)
                part_has_ingresso = await db.participantes.find_one({'_id': existing_part.get('_id'), 'ingressos.evento_id': evento_id})
                if part_has_ingresso:
                    row_errors.append('CPF já inscrito neste evento')
                else:
                    # Fallback: verifica coleção antiga ingressos_emitidos
                    ingresso_exist = await db.ingressos_emitidos.find_one({'evento_id': evento_id, 'participante_id': str(existing_part.get('_id'))})
                    if ingresso_exist:
                        row_errors.append('CPF já inscrito neste evento')

        if row_errors:
            errors.append({'line': line_no, 'errors': row_errors, 'row': row})
            continue

        # Passed validations -> if not in validate-only mode, create participant if needed and ingresso
        if not validate_only:
            if cpf_digits:
                existing_part = await db.participantes.find_one({'cpf': cpf_digits})
            else:
                existing_part = None

            if existing_part:
                participante_id = str(existing_part.get('_id'))
                reused_participants += 1
            else:
                # Extrair telefone e empresa do row, podem vir como Long do Excel
                telefone_raw = row.get('Telefone', '') or row.get('telefone', '')
                empresa_raw = row.get('Empresa', '') or row.get('empresa', '')
                
                part_doc = {
                    'nome': nome or '',
                    'email': email or '',
                    'cpf': cpf_digits or '',
                    'telefone': telefone_raw,
                    'empresa': empresa_raw
                }
                # Normalizar dados antes de inserir (converte Long->str, ''->None, etc)
                part_doc = normalize_participante_data(part_doc)
                
                res = await db.participantes.insert_one(part_doc)
                participante_id = str(res.inserted_id)
                created_participants += 1

            # tipo to use
            if tipo_obj:
                tipo_id = str(tipo_obj.get('_id') or tipo_obj.get('id') or tipo_obj.get('numero'))
            else:
                # find padrao
                tipo_obj = await db.tipos_ingresso.find_one({'evento_id': evento_id, 'padrao': True})
                tipo_id = str(tipo_obj.get('_id')) if tipo_obj else None

            ingresso_doc = {
                'evento_id': evento_id,
                'tipo_ingresso_id': tipo_id,
                'participante_id': participante_id,
                'participante_cpf': cpf_digits,
                'status': 'Ativo',
                'qrcode_hash': f'auto-{datetime.now(timezone.utc).timestamp()}',
                'data_emissao': datetime.now(timezone.utc)
            }
            # embed layout into ingresso
            layout_source = evento.get("layout_ingresso")
            try:
                from app.utils.layouts import embed_layout
                embedded = embed_layout(layout_source, {'nome': nome or '', 'cpf': cpf_digits or '', 'email': email or ''}, tipo_obj or {}, evento, ingresso_doc)
                ingresso_doc['layout_ingresso'] = embedded
            except Exception:
                pass
            # Primeiro insere na coleção antiga para compatibilidade e obter _id
            try:
                res = await db.ingressos_emitidos.insert_one(ingresso_doc)
                ingresso_doc["_id"] = res.inserted_id
            except Exception:
                # Se falhar (ex: coleção não existe), continua sem _id
                pass

            # Normalizar ingresso_doc antes de embedar (converter ObjectId->str, etc)
            ingresso_doc = normalize_participante_data(ingresso_doc)

            # Em seguida push no participante (ingressos embutidos)
            try:
                await db.participantes.update_one({"_id": ObjectId(participante_id)}, {"$push": {"ingressos": ingresso_doc}})
            except AttributeError:
                # FakeCollection in some tests lacks update_one: perform in-place append on the found document
                try:
                    part = await db.participantes.find_one({"_id": ObjectId(participante_id)})
                except Exception:
                    part = await db.participantes.find_one({"_id": participante_id})
                if part is not None:
                    part.setdefault('ingressos', []).append(ingresso_doc)
            except Exception:
                # fallback se participante._id estiver em formato string or other DB differences
                try:
                    await db.participantes.update_one({"_id": participante_id}, {"$push": {"ingressos": ingresso_doc}})
                except AttributeError:
                    part = await db.participantes.find_one({"_id": participante_id})
                    if part is not None:
                        part.setdefault('ingressos', []).append(ingresso_doc)

            created_ingressos += 1
        else:
            # In validation-only mode we don't create participants/ingressos; counts remain zero
            pass

    report = {
        'total': total,
        'created_participants': created_participants,
        'reused_participants': reused_participants,
        'created_ingressos': created_ingressos,
        'errors': errors
    }

    return report


# helper to allow awaiting in tests if needed
async def _aiter(gen):
    for x in gen:
        yield x


# --- Template generation for event (stylized) ---
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

async def generate_template_for_evento(evento: dict):
    """Gera uma planilha modelo estilizada para o evento e salva em app/static/planilhas/{nome_normalizado}_modelo.xlsx"""
    nome = evento.get('nome', 'evento')
    nome_norm = evento.get('nome_normalizado') or normalize_filename(nome)
    campos = evento.get('campos_obrigatorios_planilha', []) or []
    # Garantir os campos obrigatórios básicos
    obrigatorios = ['Nome', 'Email', 'CPF']
    for r in obrigatorios:
        if r not in campos:
            campos.insert(0, r)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Modelo'

    # Estilos simples
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Append headers
    ws.append(campos)
    for col_idx, _ in enumerate(campos, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 25

    # Add an example row with placeholder values
    example = []
    for h in campos:
        if h.lower() == 'cpf':
            example.append('123.456.789-09')
        else:
            example.append('')
    ws.append(example)

    # Instrucao sheet
    instr = wb.create_sheet('Instrucao')
    instr.append(['Instrucoes'])
    instr.append(['Preencha as colunas obrigatorias. Tipo de ingresso deve ser o numero do tipo.'])

    # Salvar arquivo
    out_dir = os.path.join('app', 'static', 'planilhas')
    os.makedirs(out_dir, exist_ok=True)
    filename = f"{nome_norm}_modelo.xlsx"
    path = os.path.join(out_dir, filename)
    wb.save(path)
    return path


def normalize_filename(name: str) -> str:
    import unicodedata, re
    nkfd = unicodedata.normalize('NFKD', name)
    ascii_str = ''.join([c for c in nkfd if not unicodedata.combining(c)])
    cleaned = re.sub(r'[^0-9a-zA-Z]', '', ascii_str)
    return cleaned.lower()


async def _fetch_evento(db, evento_id: str):
    """Busca evento tratando IDs como ObjectId ou string."""
    try:
        object_id = ObjectId(evento_id)
    except Exception:
        object_id = None

    evento = None
    if object_id:
        evento = await db.eventos.find_one({"_id": object_id})
    if not evento:
        evento = await db.eventos.find_one({"_id": evento_id})
    return evento
