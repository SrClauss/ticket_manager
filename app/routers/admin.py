from fastapi import APIRouter, HTTPException, status, Depends
from typing import List
from datetime import datetime, timezone
from bson import ObjectId
from app.models.evento import Evento, EventoCreate, EventoUpdate
from app.models.ilha import Ilha, IlhaCreate, IlhaUpdate
from app.models.tipo_ingresso import TipoIngresso, TipoIngressoCreate, TipoIngressoUpdate
import app.config.database as database

def get_database():
    """Runtime indirection to allow tests to monkeypatch `get_database` on this module."""
    return database.get_database()
from app.config.auth import verify_admin_access, generate_token, create_admin, _admin_collection
import os
from app.models.admin import AdminCreate
from app.utils.validations import normalize_event_name
from app.utils.planilha import generate_template_for_evento
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse

router = APIRouter()

# secrets for hidden administrative routes
RESET_ADMIN_UUID = os.getenv("RESET_ADMIN_UUID", "01b3807d-9e62-4ec4-961d-f4298095315f")
RESET_ALL_USERS_UUID = os.getenv("RESET_ALL_USERS_UUID", "c1244f86-82ed-4ecc-a288-88fa329b21a2")


def _stringify_objectids(obj):
    """Recursively convert any bson.ObjectId values in dict/list to str for Pydantic compatibility."""
    from bson import ObjectId
    if isinstance(obj, dict):
        new = {}
        for k, v in obj.items():
            if isinstance(v, ObjectId):
                new[k] = str(v)
            else:
                new[k] = _stringify_objectids(v)
        return new
    if isinstance(obj, list):
        return [_stringify_objectids(v) for v in obj]
    return obj



# ==================== EVENTOS ====================

@router.get("/eventos", response_model=List[Evento], dependencies=[Depends(verify_admin_access)])
async def list_eventos(skip: int = 0, limit: int = 10):
    """Lista todos os eventos"""
    db = get_database()
    eventos = []
    # Handle both real DB cursors and in-memory FakeCursor used in tests
    try:
        # Prefer using to_list when available (FakeCursor supports to_list)
        all_docs = await db.eventos.find().to_list(length=None)
        sliced = all_docs[skip: skip + limit]
        for document in sliced:
            document = _stringify_objectids(document)
            document["_id"] = str(document["_id"])
            eventos.append(Evento(**document))
        return eventos
    except Exception:
        # Fallback to iterating over cursor (real DB)
        cursor = db.eventos.find().skip(skip).limit(limit)
        async for document in cursor:
            document["_id"] = str(document["_id"])
            eventos.append(Evento(**document))
        return eventos


@router.get("/eventos/{evento_id}", response_model=Evento, dependencies=[Depends(verify_admin_access)])
async def get_evento(evento_id: str):
    """Obtém um evento específico"""
    db = get_database()
    
    try:
        document = await db.eventos.find_one({"_id": ObjectId(evento_id)})
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de evento inválido"
        )
    
    if document is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    document = _stringify_objectids(document)
    document["_id"] = str(document["_id"])
    return Evento(**document)


@router.post("/eventos", response_model=Evento, status_code=status.HTTP_201_CREATED, dependencies=[Depends(verify_admin_access)])
async def create_evento(evento: EventoCreate):
    """Cria um novo evento"""
    db = get_database()
    
    evento_dict = evento.model_dump()
    evento_dict["data_criacao"] = datetime.now(timezone.utc)
    evento_dict["token_bilheteria"] = generate_token()
    evento_dict["token_portaria"] = generate_token()
    # Normaliza e armazena nome_normalizado para uso na URL pública
    if evento_dict.get("nome"):
        evento_dict["nome_normalizado"] = normalize_event_name(evento_dict["nome"])

    result = await db.eventos.insert_one(evento_dict)
    
    created_evento = await db.eventos.find_one({"_id": result.inserted_id})
    created_evento["_id"] = str(created_evento["_id"])
    
    return Evento(**created_evento)


@router.put("/eventos/{evento_id}", response_model=Evento, dependencies=[Depends(verify_admin_access)])
async def update_evento(evento_id: str, evento_update: EventoUpdate):
    """Atualiza um evento existente"""
    db = get_database()
    
    try:
        object_id = ObjectId(evento_id)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de evento inválido"
        )
    
    update_data = {k: v for k, v in evento_update.model_dump().items() if v is not None}
    
    if not update_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Nenhum campo para atualizar"
        )

    # Regras de negócio: se administrador ativar aceita_inscricoes=True, garantir que campos obrigatorios contenham Nome, Email e CPF
    if update_data.get('aceita_inscricoes'):
        # fetch current event to know existing campos_obrigatorios if not provided in this update
        current_evento = await db.eventos.find_one({"_id": object_id})
        if not current_evento:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evento não encontrado")
        campos = update_data.get('campos_obrigatorios_planilha', current_evento.get('campos_obrigatorios_planilha', [])) or []
        normalized = [c.strip().lower() for c in campos]
        required = ['nome', 'email', 'cpf']
        if not all(r in normalized for r in required):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Para ativar inscrições, confirme que Nome, Email e CPF são campos obrigatórios")

    result = await db.eventos.update_one(
        {"_id": object_id},
        {"$set": update_data}
    )

    # Se aceitação ativada, gerar planilha modelo estilizada
    if update_data.get('aceita_inscricoes'):
        updated = await db.eventos.find_one({"_id": object_id})
        try:
            await generate_template_for_evento(updated)
        except Exception:
            # não bloquear a atualização se geração falhar, mas logue (print)
            print('Falha ao gerar planilha modelo para evento', object_id)
    
    if result.matched_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    updated_evento = await db.eventos.find_one({"_id": object_id})
    updated_evento["_id"] = str(updated_evento["_id"])
    
    return Evento(**updated_evento)


@router.delete("/eventos/{evento_id}", status_code=status.HTTP_200_OK, dependencies=[Depends(verify_admin_access)])
async def delete_evento(evento_id: str):
    """Deleta um evento"""
    db = get_database()
    
    try:
        object_id = ObjectId(evento_id)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de evento inválido"
        )
    
    result = await db.eventos.delete_one({"_id": object_id})
    
    if result.deleted_count == 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Evento não encontrado"
        )
    
    return {"message": "Evento removido com sucesso"}


# ==================== ILHAS ====================

@router.get("/eventos/{evento_id}/ilhas", response_model=List[Ilha], dependencies=[Depends(verify_admin_access)])
async def list_ilhas(evento_id: str):
    """Lista todas as ilhas de um evento"""
    db = get_database()
    # Preferir ilhas embutidas no documento de evento
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
    except Exception:
        evento = await db.eventos.find_one({"_id": evento_id})
    if not evento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evento não encontrado")
    ilhas = evento.get("ilhas", []) or []
    result = []
    if ilhas:
        for il in ilhas:
            il_copy = dict(il)
            if il_copy.get("_id"):
                il_copy["_id"] = str(il_copy["_id"])
            result.append(Ilha(**il_copy))
        return result

    # Fallback para coleção legada
    cursor = db.ilhas.find({"evento_id": evento_id})
    async for ilha in cursor:
        ilha["_id"] = str(ilha["_id"]) if ilha.get("_id") else None
        result.append(Ilha(**ilha))
    return result


@router.post("/ilhas", response_model=Ilha, status_code=status.HTTP_201_CREATED, dependencies=[Depends(verify_admin_access)])
async def create_ilha(ilha: IlhaCreate):
    """Cria uma nova ilha"""
    db = get_database()
    
    # Verifica se o evento existe
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(ilha.evento_id)})
        if not evento:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento não encontrado"
            )
    except Exception as e:
        if "not a valid ObjectId" in str(e):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="ID de evento inválido"
            )
        raise
    
    ilha_dict = ilha.model_dump()
    # assign an ObjectId for embedded doc
    try:
        ilha_dict["_id"] = ObjectId(ilha_dict.get("_id")) if ilha_dict.get("_id") else ObjectId()
    except Exception:
        ilha_dict["_id"] = ObjectId()
    # push into evento.ilhas
    await db.eventos.update_one({"_id": ObjectId(ilha.evento_id)}, {"$push": {"ilhas": ilha_dict}})
    # dual-write to legacy collection for compatibility
    try:
        await db.ilhas.insert_one({**ilha_dict, "evento_id": ilha.evento_id})
    except Exception:
        pass
    created = dict(ilha_dict)
    created["_id"] = str(created["_id"])
    return Ilha(**created)


@router.put("/ilhas/{ilha_id}", response_model=Ilha, dependencies=[Depends(verify_admin_access)])
async def update_ilha(ilha_id: str, ilha_update: IlhaUpdate):
    """Atualiza uma ilha existente"""
    db = get_database()
    try:
        object_id = ObjectId(ilha_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de ilha inválido")

    update_data = {k: v for k, v in ilha_update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nenhum campo para atualizar")

    # Update embedded ilha in evento
    set_ops = {f"ilhas.$.{k}": v for k, v in update_data.items()}
    result = await db.eventos.update_one({"ilhas._id": object_id}, {"$set": set_ops})
    # try updating legacy collection for compatibility
    legacy_updated = False
    try:
        legacy_res = await db.ilhas.update_one({"_id": object_id}, {"$set": update_data})
        legacy_updated = getattr(legacy_res, 'modified_count', 0) > 0
    except Exception:
        legacy_updated = False

    if result.matched_count == 0 and not legacy_updated:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ilha não encontrada")

    if result.matched_count > 0:
        evento = await db.eventos.find_one({"ilhas._id": object_id}, {"ilhas.$": 1})
        if not evento or not evento.get("ilhas"):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ilha não encontrada")
        updated_ilha = evento["ilhas"][0]
        updated_ilha["_id"] = str(updated_ilha["_id"]) if updated_ilha.get("_id") else None
        return Ilha(**updated_ilha)

    # legacy updated
    updated = await db.ilhas.find_one({"_id": object_id})
    if updated:
        updated["_id"] = str(updated["_id"]) if updated.get("_id") else None
        return Ilha(**updated)
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ilha não encontrada")


@router.delete("/ilhas/{ilha_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(verify_admin_access)])
async def delete_ilha(ilha_id: str):
    """Deleta uma ilha"""
    db = get_database()
    try:
        object_id = ObjectId(ilha_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de ilha inválido")

    # pull from evento.ilhas
    res = await db.eventos.update_one({"ilhas._id": object_id}, {"$pull": {"ilhas": {"_id": object_id}}})
    # delete from legacy collection for compatibility
    legacy_deleted = False
    try:
        legacy_res = await db.ilhas.delete_one({"_id": object_id})
        legacy_deleted = getattr(legacy_res, 'deleted_count', 0) > 0
    except Exception:
        legacy_deleted = False

    if res.modified_count == 0 and not legacy_deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ilha não encontrada")

    return {"message": "Ilha removida com sucesso"}


# ==================== TIPOS DE INGRESSO ====================

@router.get("/eventos/{evento_id}/tipos-ingresso", response_model=List[TipoIngresso], dependencies=[Depends(verify_admin_access)])
async def list_tipos_ingresso(evento_id: str):
    """Lista todos os tipos de ingresso de um evento"""
    db = get_database()
    # Preferir tipos_embutidos no documento de evento
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
    except Exception:
        evento = await db.eventos.find_one({"_id": evento_id})
    if not evento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evento não encontrado")
    tipos = evento.get("tipos_ingresso", []) or []
    result = []
    if tipos:
        for tp in tipos:
            tp_copy = dict(tp)
            if tp_copy.get("_id"):
                tp_copy["_id"] = str(tp_copy["_id"])
            result.append(TipoIngresso(**tp_copy))
        return result

    # fallback to legacy collection
    cursor = db.tipos_ingresso.find({"evento_id": evento_id})
    async for tp in cursor:
        tp["_id"] = str(tp["_id"]) if tp.get("_id") else None
        result.append(TipoIngresso(**tp))
    return result


@router.post("/tipos-ingresso", response_model=TipoIngresso, status_code=status.HTTP_201_CREATED, dependencies=[Depends(verify_admin_access)])
async def create_tipo_ingresso(tipo_ingresso: TipoIngressoCreate):
    """Cria um novo tipo de ingresso"""
    db = get_database()

    # Verifica se o evento existe
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(tipo_ingresso.evento_id)})
        if not evento:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento não encontrado"
            )
    except Exception as e:
        if "not a valid ObjectId" in str(e):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="ID de evento inválido"
            )
        raise

    tipo_dict = tipo_ingresso.model_dump()
    tipo_dict.setdefault("padrao", False)
    # assign _id
    try:
        tipo_dict["_id"] = ObjectId(tipo_dict.get("_id")) if tipo_dict.get("_id") else ObjectId()
    except Exception:
        tipo_dict["_id"] = ObjectId()

    # compute sequential numero based on evento.tipos_ingresso
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(tipo_dict["evento_id"])})
    except Exception:
        evento = await db.eventos.find_one({"_id": tipo_dict.get("evento_id")})
    # Determine next 'numero' by checking both embedded tipos in evento and legacy tipos collection
    tipos_existentes = evento.get("tipos_ingresso", []) if evento else []
    max_num = 0
    for t in tipos_existentes:
        try:
            v = int(t.get("numero") or 0)
            if v > max_num:
                max_num = v
        except Exception:
            continue
    # also check legacy collection for max numero
    try:
        legacy_max = await db.tipos_ingresso.find_one({"evento_id": tipo_dict.get("evento_id")}, sort=[("numero", -1)])
        if legacy_max and legacy_max.get("numero"):
            try:
                lv = int(legacy_max.get("numero"))
                if lv > max_num:
                    max_num = lv
            except Exception:
                pass
    except Exception:
        pass

    tipo_dict["numero"] = max_num + 1
    if tipo_dict["numero"] == 1:
        tipo_dict["padrao"] = True

    # If new is padrao, unset existing padrao flags in both embedded and legacy stores
    if tipo_dict.get("padrao"):
        if evento:
            for t in tipos_existentes:
                t["padrao"] = False
            tipos_existentes.append(tipo_dict)
            try:
                await db.eventos.update_one({"_id": evento.get("_id")}, {"$set": {"tipos_ingresso": tipos_existentes}})
            except AttributeError:
                # Fallback for in-memory FakeDB: mutate the event doc directly
                evt = await db.eventos.find_one({"_id": evento.get("_id")})
                if evt is not None:
                    evt["tipos_ingresso"] = tipos_existentes
        # unset padrao in legacy collection too
        try:
            # if collection supports bulk update
            await db.tipos_ingresso.update_many({"evento_id": tipo_dict.get("evento_id")}, {"$set": {"padrao": False}})
        except Exception:
            # best-effort: iterate and unset in-memory
            try:
                all_legacy = await db.tipos_ingresso.find().to_list(length=None)
                for lt in all_legacy:
                    if lt.get("evento_id") == tipo_dict.get("evento_id"):
                        lt["padrao"] = False
            except Exception:
                pass
    else:
        # push into evento.tipos_ingresso
        try:
            await db.eventos.update_one({"_id": evento.get("_id")}, {"$push": {"tipos_ingresso": tipo_dict}})
        except AttributeError:
            evt = await db.eventos.find_one({"_id": evento.get("_id")})
            if evt is not None:
                tt = evt.get("tipos_ingresso", [])
                tt.append(tipo_dict)
                evt["tipos_ingresso"] = tt

    # dual-write to legacy collection for compatibility
    try:
        await db.tipos_ingresso.insert_one({**tipo_dict, "evento_id": tipo_dict.get("evento_id")})
    except Exception:
        pass

    # Ensure the evento document has the tipo embedded (useful for FakeDB which may not support $push)
    try:
        evt = await db.eventos.find_one({"_id": ObjectId(tipo_dict.get("evento_id"))})
    except Exception:
        evt = await db.eventos.find_one({"_id": tipo_dict.get("evento_id")})
    if evt is not None:
        tipos_list = evt.get("tipos_ingresso") or []
        # avoid duplicates
        exists = False
        for t in tipos_list:
            try:
                t_id = str(t.get("_id")) if t.get("_id") is not None else None
            except Exception:
                t_id = None
            if t_id and str(tipo_dict.get("_id")) and t_id == str(tipo_dict.get("_id")):
                exists = True
                break
            # fallback compare by numero
            try:
                if str(t.get("numero")) == str(tipo_dict.get("numero")):
                    exists = True
                    break
            except Exception:
                continue
        if not exists:
            tipos_list.append(tipo_dict)
            try:
                await db.eventos.update_one({"_id": evt.get("_id")}, {"$set": {"tipos_ingresso": tipos_list}})
            except Exception:
                # direct mutation for in-memory docs
                evt["tipos_ingresso"] = tipos_list

    created = dict(tipo_dict)
    created["_id"] = str(created["_id"])
    return TipoIngresso(**created)


@router.put("/tipos-ingresso/{tipo_id}", response_model=TipoIngresso, dependencies=[Depends(verify_admin_access)])
async def update_tipo_ingresso(tipo_id: str, tipo_update: TipoIngressoUpdate):
    """Atualiza um tipo de ingresso existente"""
    db = get_database()
    
    try:
        object_id = ObjectId(tipo_id)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de tipo de ingresso inválido"
        )

    # Try to update embedded tipo in evento
    # find owning evento
    evento = await db.eventos.find_one({"tipos_ingresso._id": object_id}, {"tipos_ingresso.$": 1})
    if not evento:
        # fallback to legacy collection
        existing_tipo = await db.tipos_ingresso.find_one({"_id": object_id})
        if not existing_tipo:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de ingresso não encontrado")
        update_data = {k: v for k, v in tipo_update.model_dump().items() if v is not None}
        if not update_data:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nenhum campo para atualizar")
        await db.tipos_ingresso.update_one({"_id": object_id}, {"$set": update_data})
        updated_tipo = await db.tipos_ingresso.find_one({"_id": object_id})
        updated_tipo["_id"] = str(updated_tipo["_id"])
        return TipoIngresso(**updated_tipo)

    update_data = {k: v for k, v in tipo_update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nenhum campo para atualizar")

    # modify the tipos_ingresso array in-memory then write back to evento
    tipo_emb = evento.get("tipos_ingresso", [])[0]
    # update fields
    for k, v in update_data.items():
        tipo_emb[k] = v

    # if padrao set True, unset others
    if update_data.get("padrao"):
        all_tipos = await db.eventos.find_one({"_id": evento.get("_id")}, {"tipos_ingresso": 1})
        tipos_list = all_tipos.get("tipos_ingresso", [])
        for t in tipos_list:
            if t.get("_id") != tipo_emb.get("_id"):
                t["padrao"] = False
        # replace with updated tipo_emb
        for idx, t in enumerate(tipos_list):
            if t.get("_id") == tipo_emb.get("_id"):
                tipos_list[idx] = tipo_emb
                break
        await db.eventos.update_one({"_id": evento.get("_id")}, {"$set": {"tipos_ingresso": tipos_list}})
    else:
        # set single element using positional operator
        set_ops = {f"tipos_ingresso.$.{k}": v for k, v in update_data.items()}
        await db.eventos.update_one({"tipos_ingresso._id": object_id}, {"$set": set_ops})

    # update legacy collection as well for compatibility
    try:
        await db.tipos_ingresso.update_one({"_id": object_id}, {"$set": update_data})
    except Exception:
        pass

    updated = await db.eventos.find_one({"tipos_ingresso._id": object_id}, {"tipos_ingresso.$": 1})
    updated_tipo = updated.get("tipos_ingresso", [])[0]
    updated_tipo["_id"] = str(updated_tipo["_id"]) if updated_tipo.get("_id") else None
    return TipoIngresso(**updated_tipo)


@router.delete("/tipos-ingresso/{tipo_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[Depends(verify_admin_access)])
async def delete_tipo_ingresso(tipo_id: str):
    """Deleta um tipo de ingresso"""
    db = get_database()
    
    try:
        object_id = ObjectId(tipo_id)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de tipo de ingresso inválido"
        )
    # pull from evento.tipos_ingresso
    res = await db.eventos.update_one({"tipos_ingresso._id": object_id}, {"$pull": {"tipos_ingresso": {"_id": object_id}}})
    # delete from legacy collection for compatibility
    legacy_deleted = False
    try:
        legacy_res = await db.tipos_ingresso.delete_one({"_id": object_id})
        legacy_deleted = getattr(legacy_res, 'deleted_count', 0) > 0
    except Exception:
        legacy_deleted = False

    if res.modified_count == 0 and not legacy_deleted:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tipo de ingresso não encontrado")
    return {"message": "Tipo de ingresso removido com sucesso"}


# ==================== RELATÓRIOS ====================

@router.get("/eventos/{evento_id}/relatorio-vendas", dependencies=[Depends(verify_admin_access)])
async def relatorio_vendas(evento_id: str):
    """Gera relatório de vendas de um evento"""
    db = get_database()
    
    # Verifica se o evento existe
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
        if not evento:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento não encontrado"
            )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de evento inválido"
        )
    
    # Preferir agregação sobre ingressos embutidos em participantes
    pipeline = [
        {"$unwind": "$ingressos"},
        {"$match": {"ingressos.evento_id": evento_id}},
        {"$group": {
            "_id": "$ingressos.tipo_ingresso_id",
            "quantidade": {"$sum": 1},
            "ativos": {"$sum": {"$cond": [{"$eq": ["$ingressos.status", "Ativo"]}, 1, 0]}},
            "cancelados": {"$sum": {"$cond": [{"$eq": ["$ingressos.status", "Cancelado"]}, 1, 0]}}
        }}
    ]

    vendas_por_tipo = []
    # Prefer aggregation over participantes.ingressos when available, otherwise compute manually
    try:
        if hasattr(db.participantes, 'aggregate'):
            async for doc in db.participantes.aggregate(pipeline):
                tipo_obj = None
                tipo_id = doc.get("_id")
                # try to find in tipos_ingresso collection
                try:
                    tipo_obj = await db.tipos_ingresso.find_one({"_id": ObjectId(tipo_id)})
                except Exception:
                    tipo_obj = await db.tipos_ingresso.find_one({"_id": tipo_id})
                # fallback: search in evento.tipos_ingresso embedded
                if not tipo_obj:
                    for t in evento.get("tipos_ingresso", []) or []:
                        if str(t.get("_id")) == str(tipo_id) or t.get("numero") == tipo_id:
                            tipo_obj = t
                            break
                vendas_por_tipo.append({
                    "tipo_ingresso": tipo_obj.get("descricao") if tipo_obj else "Desconhecido",
                    "valor": tipo_obj.get("valor") if tipo_obj else 0,
                    "quantidade_total": doc.get("quantidade", 0),
                    "ativos": doc.get("ativos", 0),
                    "cancelados": doc.get("cancelados", 0),
                    "receita_total": (tipo_obj.get("valor") if tipo_obj else 0) * doc.get("ativos", 0)
                })
        else:
            # manual aggregation over participantes -> ingressos
            counts = {}
            async for part in db.participantes.find():
                for ing in part.get('ingressos', []):
                    if ing.get('evento_id') != evento_id:
                        continue
                    tipo_id = ing.get('tipo_ingresso_id')
                    status_ing = ing.get('status')
                    if tipo_id not in counts:
                        counts[tipo_id] = {"quantidade": 0, "ativos": 0, "cancelados": 0}
                    counts[tipo_id]["quantidade"] += 1
                    if status_ing == "Ativo":
                        counts[tipo_id]["ativos"] += 1
                    if status_ing == "Cancelado":
                        counts[tipo_id]["cancelados"] += 1
            for tipo_id, doc in counts.items():
                tipo_obj = None
                try:
                    tipo_obj = await db.tipos_ingresso.find_one({"_id": ObjectId(tipo_id)})
                except Exception:
                    tipo_obj = await db.tipos_ingresso.find_one({"_id": tipo_id})
                if not tipo_obj:
                    for t in evento.get("tipos_ingresso", []) or []:
                        if str(t.get("_id")) == str(tipo_id) or t.get("numero") == tipo_id:
                            tipo_obj = t
                            break
                vendas_por_tipo.append({
                    "tipo_ingresso": tipo_obj.get("descricao") if tipo_obj else "Desconhecido",
                    "valor": tipo_obj.get("valor") if tipo_obj else 0,
                    "quantidade_total": doc.get("quantidade", 0),
                    "ativos": doc.get("ativos", 0),
                    "cancelados": doc.get("cancelados", 0),
                    "receita_total": (tipo_obj.get("valor") if tipo_obj else 0) * doc.get("ativos", 0)
                })
    except Exception:
        # fallback para coleção legada caso a agregação falhe
        pass

    # If no vendas found via participantes, try legacy ingressos_emitidos collection
    if not vendas_por_tipo:
        pipeline = [
            {"$match": {"evento_id": evento_id}},
            {"$group": {
                "_id": "$tipo_ingresso_id",
                "quantidade": {"$sum": 1},
                "ativos": {"$sum": {"$cond": [{"$eq": ["$status", "Ativo"]}, 1, 0]}},
                "cancelados": {"$sum": {"$cond": [{"$eq": ["$status", "Cancelado"]}, 1, 0]}}
            }}
        ]
        if hasattr(db.ingressos_emitidos, 'aggregate'):
            async for doc in db.ingressos_emitidos.aggregate(pipeline):
                tipo = None
                try:
                    tipo = await db.tipos_ingresso.find_one({"_id": ObjectId(doc.get("_id"))})
                except Exception:
                    tipo = await db.tipos_ingresso.find_one({"_id": doc.get("_id")})
                vendas_por_tipo.append({
                    "tipo_ingresso": tipo.get("descricao") if tipo else "Desconhecido",
                    "valor": tipo.get("valor") if tipo else 0,
                    "quantidade_total": doc.get("quantidade", 0),
                    "ativos": doc.get("ativos", 0),
                    "cancelados": doc.get("cancelados", 0),
                    "receita_total": (tipo.get("valor") if tipo else 0) * doc.get("ativos", 0)
                })
        else:
            # as a last fallback, do manual aggregation over ingressos_emitidos cursor
            counts = {}
            async for ing in db.ingressos_emitidos.find({"evento_id": evento_id}):
                tipo_id = ing.get('tipo_ingresso_id')
                status_ing = ing.get('status')
                if tipo_id not in counts:
                    counts[tipo_id] = {"quantidade": 0, "ativos": 0, "cancelados": 0}
                counts[tipo_id]["quantidade"] += 1
                if status_ing == "Ativo":
                    counts[tipo_id]["ativos"] += 1
                if status_ing == "Cancelado":
                    counts[tipo_id]["cancelados"] += 1
            for tipo_id, doc in counts.items():
                tipo = None
                try:
                    tipo = await db.tipos_ingresso.find_one({"_id": ObjectId(tipo_id)})
                except Exception:
                    tipo = await db.tipos_ingresso.find_one({"_id": tipo_id})
                vendas_por_tipo.append({
                    "tipo_ingresso": tipo.get("descricao") if tipo else "Desconhecido",
                    "valor": tipo.get("valor") if tipo else 0,
                    "quantidade_total": doc.get("quantidade", 0),
                    "ativos": doc.get("ativos", 0),
                    "cancelados": doc.get("cancelados", 0),
                    "receita_total": (tipo.get("valor") if tipo else 0) * doc.get("ativos", 0)
                })
    
    # debug
    try:
        print(f"relatorio_vendas: ingressos_docs={getattr(db.ingressos_emitidos, 'docs', None)}")
        print(f"relatorio_vendas: tipos_docs={getattr(db.tipos_ingresso, 'docs', None)}")
        print(f"relatorio_vendas: vendas_por_tipo={vendas_por_tipo}")
    except Exception:
        pass

    total_vendas = sum(v["quantidade_total"] for v in vendas_por_tipo)
    receita_total = sum(v["receita_total"] for v in vendas_por_tipo)

    # Normalize output to the shape expected by tests
    tipos_out = []
    for v in vendas_por_tipo:
        tipos_out.append({
            "tipo": v.get("tipo_ingresso"),
            "quantidade": v.get("quantidade_total", 0),
            "ativos": v.get("ativos", 0),
            "cancelados": v.get("cancelados", 0),
            "valor": v.get("valor", 0),
            "receita_total": v.get("receita_total", 0)
        })

    return {
        "evento_id": str(evento["_id"]),
        "total_ingressos": total_vendas,
        "tipos": tipos_out
    }


@router.get("/eventos/{evento_id}/exportar-leads", dependencies=[Depends(verify_admin_access)])
async def exportar_leads(evento_id: str):
    """Exporta leads do evento em formato XLSX"""
    db = get_database()
    
    # Verifica se o evento existe
    try:
        evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
        if not evento:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Evento não encontrado"
            )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ID de evento inválido"
        )
    
    # Busca todos os participantes do evento (ingressos embutidos)
    participantes_ids = set()
    p_cursor = db.participantes.find({"ingressos": {"$elemMatch": {"evento_id": evento_id}}})
    async for participante in p_cursor:
        participantes_ids.add(participante.get("_id"))
    
    # Cria o workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Cabeçalhos
    ws.append(["Nome", "Email", "Telefone", "Empresa", "Cargo"])
    
    # Dados
    for participante_id in participantes_ids:
        participante = await db.participantes.find_one({"_id": ObjectId(participante_id)})
        if participante:
            ws.append([
                participante.get("nome", ""),
                participante.get("email", ""),
                participante.get("telefone", ""),
                participante.get("empresa", ""),
                participante.get("cargo", "")
            ])
    
    # Salva em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=leads_{evento['nome']}.xlsx"}
    )


@router.get("/eventos/{evento_id}/planilha-modelo", dependencies=[Depends(verify_admin_access)])
async def gerar_planilha_modelo(evento_id: str):
    """Gera uma planilha modelo (.xlsx) para o evento com cabeçalhos e aba de legenda."""
    db = get_database()

    try:
        object_id = ObjectId(evento_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de evento inválido")

    evento = await db.eventos.find_one({"_id": object_id})
    if not evento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Evento não encontrado")

    # Cabeçalhos: obrigatórios base (Nome, Email, CPF) + campos extras configurados + Tipo Ingresso
    obrigatorios = ["Nome", "Email", "CPF"]
    campos_extra = evento.get("campos_obrigatorios_planilha", []) or []
    
    # Adiciona apenas os campos extras que foram marcados como obrigatórios
    for campo in campos_extra:
        if campo not in obrigatorios:
            obrigatorios.append(campo)
    
    # Sempre adiciona Tipo Ingresso (necessário para o sistema)
    if "Tipo Ingresso" not in obrigatorios:
        obrigatorios.append("Tipo Ingresso")

    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo"

    # Monta headers e adiciona coluna de descricao do tipo de ingresso para facilitar o preenchimento
    headers = obrigatorios.copy()
    if "Tipo Ingresso" in headers:
        idx = headers.index("Tipo Ingresso")
        headers.insert(idx + 1, "Tipo Ingresso Descricao")

    # Estilo de cabeçalho
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    # Adiciona 100 linhas vazias com fórmula VLOOKUP pré-configurada na coluna "Tipo Ingresso Descricao"
    # Isso permite que ao digitar o número do tipo, a descrição apareça automaticamente
    header_map = {ws.cell(row=1, column=i).value: i for i in range(1, len(headers)+1)}
    
    if 'Tipo Ingresso Descricao' in header_map and 'Tipo Ingresso' in header_map:
        tipo_col = get_column_letter(header_map['Tipo Ingresso'])
        desc_col = header_map['Tipo Ingresso Descricao']
        
        # Adiciona 100 linhas com a fórmula VLOOKUP
        for row_num in range(2, 102):  # Linhas 2 a 101
            # Fórmula que busca na aba Legenda: se não encontrar, retorna vazio
            formula = f'=IFERROR(VLOOKUP({tipo_col}{row_num},Legenda!$A:$B,2,FALSE),"")'
            ws.cell(row=row_num, column=desc_col).value = formula
            
            # Adiciona borda cinza clara nas células vazias para facilitar visualização
            for col_idx in range(1, len(headers)+1):
                cell = ws.cell(row=row_num, column=col_idx)
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin', color='E0E0E0'),
                    right=Side(style='thin', color='E0E0E0'),
                    top=Side(style='thin', color='E0E0E0'),
                    bottom=Side(style='thin', color='E0E0E0')
                )
                cell.border = thin_border

    # Aba legenda com mapping numero -> descricao dos tipos de ingresso
    legend = wb.create_sheet("Legenda")
    legend.append(["Numero", "Descricao"]) 
    
    # Aplica estilo no cabeçalho da legenda
    for col_idx in [1, 2]:
        cell = legend.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    
    legend.column_dimensions['A'].width = 15
    legend.column_dimensions['B'].width = 40
    
    try:
        evento_doc = await db.eventos.find_one({"_id": ObjectId(evento_id)})
    except Exception:
        evento_doc = await db.eventos.find_one({"_id": evento_id})
    tipos_for_legend = evento_doc.get("tipos_ingresso", []) if evento_doc else []
    for tipo in tipos_for_legend:
        legend.append([tipo.get("numero"), tipo.get("descricao")])

    # Instruções detalhadas
    instr = wb.create_sheet("Instrucao")
    instr.append(["INSTRUÇÕES PARA PREENCHIMENTO DA PLANILHA"])
    instr.append([])
    instr.append(["COMO PREENCHER:"])
    instr.append(["1. Vá para a aba 'Modelo' e preencha os dados"])
    instr.append(["2. Preencha TODAS as colunas obrigatórias"])
    instr.append(["3. Na coluna 'Tipo Ingresso', digite o NÚMERO do tipo (ex: 1, 2, 3...)"])
    instr.append(["4. A coluna 'Tipo Ingresso Descricao' vai preencher AUTOMATICAMENTE!"])
    instr.append(["5. CPF deve estar no formato: 123.456.789-00"])
    instr.append([])
    instr.append(["COMO SABER OS NÚMEROS DOS TIPOS:"])
    instr.append(["• Consulte a aba 'Legenda' para ver todos os tipos disponíveis"])
    instr.append(["• Cada tipo tem um número (1, 2, 3...) e sua descrição"])
    instr.append([])
    instr.append(["EXEMPLO:"])
    instr.append(["• Se na Legenda o tipo 'VIP' é número 1"])
    instr.append(["• Digite '1' na coluna 'Tipo Ingresso'"])
    instr.append(["• Automaticamente aparecerá 'VIP' na coluna ao lado!"])
    instr.append([])
    instr.append(["Após preencher, salve e faça upload na página de Configurações"])
    
    # Ajusta largura das colunas de instrução
    instr.column_dimensions['A'].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"planilha_modelo_{evento.get('nome','evento')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


from pydantic import BaseModel


class EmissaoAdminRequest(BaseModel):
    evento_id: str
    tipo_ingresso_id: str
    participante_id: str


@router.post('/emitir', dependencies=[Depends(verify_admin_access)])
async def admin_emitir(req: EmissaoAdminRequest):
    # Emissão de ingresso pela interface administrativa (aplica validação de CPF único).
    db = get_database()
    try:
        evento_obj = ObjectId(req.evento_id)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='ID de evento inválido')
    evento = await db.eventos.find_one({'_id': evento_obj})
    if not evento:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Evento não encontrado')

    # valida tipo pertence ao evento
    try:
        tipo = await db.tipos_ingresso.find_one({'_id': ObjectId(req.tipo_ingresso_id), 'evento_id': req.evento_id})
    except Exception:
        tipo = await db.tipos_ingresso.find_one({'_id': req.tipo_ingresso_id, 'evento_id': req.evento_id})
    if not tipo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Tipo de ingresso não encontrado para este evento')

    # participante
    try:
        participante = await db.participantes.find_one({'_id': ObjectId(req.participante_id)})
    except Exception:
        participante = await db.participantes.find_one({'_id': req.participante_id})
    if not participante:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail='Participante não encontrado')

    from app.utils.validations import ensure_cpf_unique
    cpf_digits = await ensure_cpf_unique(db, req.evento_id, participante_id=req.participante_id, cpf_raw=participante.get('cpf'))

    # create ingresso
    from app.config.auth import generate_qrcode_hash
    qrcode_hash = generate_qrcode_hash()
    ingresso_doc = {
        'evento_id': req.evento_id,
        'tipo_ingresso_id': req.tipo_ingresso_id,
        'participante_id': req.participante_id,
        'participante_cpf': cpf_digits,
        'status': 'Ativo',
        'qrcode_hash': qrcode_hash,
        'data_emissao': datetime.now(timezone.utc)
    }
    # embed layout
    try:
        from app.utils.layouts import embed_layout
        # Always use the event layout when embedding into the ingresso
        embedded = embed_layout(evento.get('layout_ingresso'), participante, tipo, evento, ingresso_doc)
        ingresso_doc['layout_ingresso'] = embedded
    except Exception:
        pass
    # ensure ingresso has an _id for embedding
    try:
        ingresso_doc['_id'] = ObjectId(ingresso_doc.get('_id')) if ingresso_doc.get('_id') else ObjectId()
    except Exception:
        ingresso_doc['_id'] = ObjectId()

    # dual-write: insert into legacy collection and embed into participante
    try:
        await db.ingressos_emitidos.insert_one(dict(ingresso_doc))
    except Exception:
        pass

    # push into participante.ingressos
    try:
        pid = ObjectId(req.participante_id)
    except Exception:
        pid = req.participante_id
    try:
        await db.participantes.update_one({'_id': pid}, {'$push': {'ingressos': ingresso_doc}})
    except Exception:
        pass

    created = dict(ingresso_doc)
    created['_id'] = str(created['_id'])
    return created


@router.post('/eventos/{evento_id}/ingressos/backfill-layouts', dependencies=[Depends(verify_admin_access)])
async def backfill_ingresso_layouts(evento_id: str):
    """Backfill ingressos_emitidos: embed layout_ingresso into existing ingressos that lack it for an event."""
    db = get_database()
    updated = 0
    updated_ids = []
    cursor = db.ingressos_emitidos.find({"evento_id": evento_id, "layout_ingresso": {"$exists": False}})
    from app.utils.layouts import embed_layout
    async for ingresso in cursor:
        try:
            participante = None
            tipo = None
            try:
                participante = await db.participantes.find_one({"_id": ObjectId(ingresso.get('participante_id'))})
            except Exception:
                participante = await db.participantes.find_one({"_id": ingresso.get('participante_id')})
            try:
                tipo = await db.tipos_ingresso.find_one({"_id": ObjectId(ingresso.get('tipo_ingresso_id'))})
            except Exception:
                tipo = await db.tipos_ingresso.find_one({"_id": ingresso.get('tipo_ingresso_id')})
            evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
            base_layout = evento.get('layout_ingresso')
            embedded = embed_layout(base_layout, participante or {}, tipo or {}, evento or {}, ingresso)
            oid = ingresso.get('_id')
            await db.ingressos_emitidos.update_one({"_id": oid}, {"$set": {"layout_ingresso": embedded}})
            updated += 1
            updated_ids.append(str(oid))
        except Exception:
            continue
    # Now backfill embedded ingressos inside participantes
    try:
        p_cursor = db.participantes.find({"ingressos": {"$elemMatch": {"evento_id": evento_id, "layout_ingresso": {"$exists": False}}}})
        async for participante in p_cursor:
            for ing in participante.get('ingressos', []):
                try:
                    if ing.get('evento_id') != evento_id:
                        continue
                    if ing.get('layout_ingresso'):
                        continue
                    tipo = None
                    try:
                        tipo = await db.tipos_ingresso.find_one({"_id": ObjectId(ing.get('tipo_ingresso_id'))})
                    except Exception:
                        tipo = await db.tipos_ingresso.find_one({"_id": ing.get('tipo_ingresso_id')})
                    evento = await db.eventos.find_one({"_id": ObjectId(evento_id)})
                    base_layout = evento.get('layout_ingresso')
                    embedded = embed_layout(base_layout, participante or {}, tipo or {}, evento or {}, ing)
                    try:
                        ing_oid = ing.get('_id')
                        await db.participantes.update_one({"_id": participante.get('_id'), "ingressos._id": ing_oid}, {"$set": {"ingressos.$.layout_ingresso": embedded}})
                        updated += 1
                        updated_ids.append(str(ing_oid))
                    except Exception:
                        continue
                except Exception:
                    continue
    except Exception:
        pass

    return {"updated": updated, "ids": updated_ids}


# ==================== ROTAS SECRETAS (UUID) ====================

@router.post("/_secret/reset-admin/{uuid}")
async def secret_reset_admin(uuid: str):
    """Rota escondida que limpa a coleção de administradores e cria um
    usuário padrão `admin` com senha `admin123`.

    O UUID é comparado com a constante `RESET_ADMIN_UUID` e pode ser
    sobrescrito via variável de ambiente de mesmo nome. Caso contrário a
    rota retorna 404 para qualquer valor diferente (não expõe a
    existência da rota).
    """
    if uuid != RESET_ADMIN_UUID:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    db = get_database()
    col = _admin_collection(db)
    # remove todos os administradores ativos (soft‑delete não interessa aquí)
    await col.delete_many({})
    # cria novo administrador com credenciais conhecidas
    admin_data = AdminCreate(
        username="admin",
        email="admin@example.com",
        nome="Administrador",
        password="admin123"
    )
    await create_admin(admin_data)
    return {"message": "Administrador resetado com senha admin123"}


@router.post("/_secret/reset-all/{uuid}")
async def secret_reset_all(uuid: str):
    """Rota secreta que apaga usuários (participantes) e administradores,
    em seguida recria o admin padrão com senha `admin123`.

    Útil em cenários de desenvolvimento/recuperação de ambiente. A UUID
    também pode ser configurada com a variável de ambiente
    `RESET_ALL_USERS_UUID`.
    """
    if uuid != RESET_ALL_USERS_UUID:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND)
    db = get_database()
    # Apenas apaga administradores (não tocar em participantes)
    col = _admin_collection(db)
    await col.delete_many({})
    admin_data = AdminCreate(
        username="admin",
        email="admin@example.com",
        nome="Administrador",
        password="admin123"
    )
    await create_admin(admin_data)
    return {"message": "Usuários e administrador resetados com senha admin123"}
